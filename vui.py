import tkinter
import customtkinter
import tkinter.messagebox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
import time
import re
import serial
import minimalmodbus



##########################
# Evan Peck
# Capstone Project
# 4/11/23
# Group 6, Skittlz Dynasty
# GUI for control of Generant Compaction Station
##########################

###### NOTES ########
# Add AutoLiv and UofU images
# Get feedback on color pallet?
# Add opening window w/ images, manual/automatic
# Find frequency of rotation direction change
# Error message for invalid filename
# Fix errors in filesave if not all subsystems are active

customtkinter.set_appearance_mode("dark")           #Set overall apearance of window, can change to "light"
customtkinter.set_default_color_theme("blue")

#Safety bounds and subsystem limitations
#Primary concern is high frequency (>150Hz) with a 1mm amplitude. Max motor torque may be exceeded.
freqBounds1 = [20, 50]                         #Set limitations on allowable frequency [Hz]
freqBounds06 = [50, 125]
freqBounds04 = [50, 150]
freqBoundsCustom = [50, 150]
freqBounds = [freqBounds1[0], freqBounds1[1]]

forceBounds1 = [0, 300]                         #Set limitations on allowable compaction force [N]
forceBounds06 = [0, 400]
forceBounds04 = [0, 500]
forceBoundsCustom = [0, 550]
forceBounds = [forceBounds1[0], forceBounds1[1]]

rotSpeedBounds1 = [0, 1]                       #Set limitations on allowable rotation oscillation frequecy [Hz]
rotSpeedBounds06 = [0, 1]
rotSpeedBounds04 = [0, 1]
rotSpeedBoundsCustom = [0, 1]
rotSpeedBounds = [rotSpeedBounds1[0], rotSpeedBounds1[1]]

timeBounds = [0, 60]                            #Allowable test run time

freqResult = 75             #Temporary results used for testing ***********************************************
forceResult = 250
rotateResult = 9.81
timeResult = 5
global clearcore
clearcoreResults = None                 #Initialize communication with ClearCore board for serial communication
clearcore = serial.Serial('COM3', 9600, timeout=0.5)


#Temp measurements
t = np.arange(0.0, 3.0, 0.01)                   #Temporary results generation used for testing
freqMeasures = np.sin(np.pi * t) + 80
forceMeasures = np.cos(np.pi * t)+150
rotationMeasures = 9.8*np.sin(np.pi * t*5)
maxFreq = max(freqMeasures)
maxForce = max(forceMeasures)
maxRotation = max(rotationMeasures)
maxTime = max(t)


class ResultsTab(customtkinter.CTkTabview):                 #Display results in a tab view within a new window generated in ProcessResults
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        global t, freqMeasures, forceMeasures, rotationMeasures

        #Create tabs
        self.add("Measured Results")
        self.add("Freq. Results")
        self.add("Force Results")
        self.add("Rot. Speed Results")

        #Insert results in tabs
        measured_freq = 'Measured frequency:               ' + str(freqResult) + 'Hz\n\n'
        measured_force = 'Measured force:                        ' + str(forceResult) + 'N\n\n'
        measured_rotation = "Measured rotation frequency:       " + str(rotateResult) + "Hz\n\n"
        measured_time = "Measured run time:                   " + str(timeResult) + "sec."
        self.measuredText = customtkinter.CTkTextbox(master=self.tab("Measured Results"))
        self.measuredText.pack(fill='both', padx=10, pady=10)
        self.measuredText.insert("0.0", measured_freq + measured_force + measured_rotation + measured_time)

        # Frequency plot
        freqFig = Figure(figsize=(8, 5))  # Initialize figure
        freqPlot = freqFig.add_subplot(111)
        freqPlot.axhline(y=maxFreq, color='r', linestyle='--')                               # Horizontal line of max frequency
        freqPlot.annotate(str(maxFreq), xy=(t[-1], maxFreq), xytext=(20,0), color='r',       # Annotate the max frequency line
                          textcoords="offset points", va='center')
        freqPlot.plot(t, freqMeasures)                                                       # Plot the measurements
        freqPlot.set(xlabel='Time [sec.]', ylabel='Freq. [Hz]')

        freqCanvas = FigureCanvasTkAgg(freqFig, master=self.tab("Freq. Results"))            # Display the figure in the appropriate results tab
        freqCanvas.draw()
        freqCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)

        #Force plot
        forceFig = Figure(figsize=(5,5))
        forcePlot = forceFig.add_subplot(111)
        forcePlot.axhline(y=maxForce, color='r', linestyle='--')
        forcePlot.annotate(str(maxForce), xy=(t[-1], maxForce), xytext=(20, 0), color='r',
                          textcoords="offset points", va='center')
        forcePlot.plot(t, forceMeasures)
        forcePlot.set(xlabel='Time [sec.]', ylabel='Force [N]')

        forceCanvas = FigureCanvasTkAgg(forceFig, master=self.tab("Force Results"))
        forceCanvas.draw()
        forceCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)

        #Rotation plot
        rotationFig = Figure(figsize=(5, 5))
        rotationPlot = rotationFig.add_subplot(111)
        rotationPlot.axhline(y=maxRotation, color='r', linestyle='--')
        rotationPlot.annotate(str(maxRotation), xy=(t[-1], maxForce), xytext=(20, 0), color='r',
                           textcoords="offset points", va='center')
        rotationPlot.plot(t, rotationMeasures)
        rotationPlot.set(xlabel='Time [sec.]', ylabel='Speed [rad/s]')

        rotationCanvas = FigureCanvasTkAgg(rotationFig, master=self.tab("Rot. Speed Results"))
        rotationCanvas.draw()
        rotationCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)



class ProcessResults(customtkinter.CTkToplevel):            # Create new window to display test results in
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title("Process Results")   
        self.geometry("700x500")

        self.tab_view = ResultsTab(master=self)             # Associate with results tabs created above
        self.tab_view.pack(fill='both', expand=True, padx=10, pady=10)



class App(customtkinter.CTk):                               ## Main window with subsytem controls ##
    def __init__(self):
        super().__init__()

        self.title("AutoLiv Generant Compaction Station Control")
        self.geometry(f"{1100}x{600}")

        # Grid Layout
        self.grid_columnconfigure((1, 2), weight=0)
        self.grid_rowconfigure((0, 1, 2, 3), weight=0)

        ## Radiobuttons for cam amplitude selection ##
        self.radiobutton_frame = customtkinter.CTkFrame(self)
        self.radiobutton_frame.grid(row=0, column=0, rowspan=1, columnspan=1, padx=(10, 0), pady=(20, 0), sticky="nsew")
        self.radio_var = tkinter.IntVar(value=0)
        self.label_radio_group = customtkinter.CTkLabel(master=self.radiobutton_frame, text="Cam Amplitude", font=('CTkFont', 15))
        self.label_radio_group.grid(row=0, column=1, columnspan=2, padx=10, pady=(20, 20), sticky="ew")
        self.amp_1 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="1mm",
                                                  variable=self.radio_var, value=0, command=self.setBounds)
        self.amp_1.grid(row=1, column=0, padx=20, pady=10, sticky="n")
        self.amp_06 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="0.6mm",
                                                   variable=self.radio_var, value=1, command=self.setBounds)
        self.amp_06.grid(row=1, column=1, padx=20, pady=10, sticky="n")
        self.amp_04 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="0.4mm",
                                                   variable=self.radio_var, value=2, command=self.setBounds)
        self.amp_04.grid(row=1, column=2, padx=20, pady=10, sticky="n")
        self.amp_custom = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="Custom",
                                                   variable=self.radio_var, value=3, command=self.setBounds)
        self.amp_custom.grid(row=1, column=3, padx=20, pady=10, sticky="n")

        ## Frequency Slider ##
        # Subsystem on/off control, slider and text entry for frequency selection
        self.slide_frame = customtkinter.CTkFrame(self)#, fg_color="transparent")                                       # Create frame for all subsystem sliders
        self.slide_frame.grid(row=1, column=0, rowspan=4, padx=(10, 0), pady=(20, 20), sticky="nsew")                   # Place the frame within the window
        self.slide_frame.grid_columnconfigure(1, weight=1)
        self.slide_frame.grid_rowconfigure(2, weight=1)
        self.subsysLabel = customtkinter.CTkLabel(self.slide_frame, text="Subsystems Control", font=('CTkFont', 15))    # Create a label for the frame
        self.subsysLabel.grid(row=0, column=1, pady=(10,0))
        self.freq_var = "disabled"
        self.freqSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",         # Create the slider for frequency selection, begins disabled until the subsystem is activated w/ checkbox
                                                  number_of_steps=250, from_=freqBounds[0], to=freqBounds[1])           # Slider bounds can be adjusted
        self.freqSlider.bind("<ButtonRelease-1>", self.checkFreqSlider)                                                 # Call to check if slider matches entry (if not, change entry to slider value)
        self.freqSlider.grid(row=1, column=1, columnspan=2, padx=(10, 10), pady=(20, 30), sticky="ew")                  # Place the slider within the frame

        self.vibButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nVibration",                          # Create checkbox for frequency subsystem activation
                                                   command=self.freqToggle, onvalue=1, offvalue=0)                      # Bound to command which enables/disables slider/entry depending on checkbox state
        self.vibButton.grid(row=1, column=0, padx=(10, 10), pady=(20, 30))                                              # Place the checkbox within the frame

        self.freqEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="Hz", state="disabled",              # Create the entry box for the frequency selection, begins disabled until the subsystem is activated w/ checkbox
                                                validate='focusout', validatecommand=self.checkFreqEntry, width=50)     # Call to check if entry box matches slider (if not, change the slider to entry value)
        self.freqEntry.grid(row=1, column=3, padx=(10, 10), pady=(20, 30), sticky='w')                                  # Place the entry box within the frame

        self.freqLabel = customtkinter.CTkLabel(self.slide_frame, text="Hz   ")                                         # Create a label for the entry box displaying the desired units
        self.freqLabel.grid(row=1, column=4, padx=(0, 10), pady=(20, 30), sticky='w')                                   # Place the label within the frame

        ## Force Slider ##
        # Subsystem on/off control, slider and text entry for compaction force selection
        # Follows the same setup as the frequency slider section above
        self.force_var = "disabled"
        self.forceSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                   number_of_steps=250, from_=forceBounds[0], to=forceBounds[1])
        self.forceSlider.bind("<ButtonRelease-1>", self.checkForceSlider)
        self.forceSlider.grid(row=2, column=1, columnspan=2, padx=(10, 10), pady=(30, 30), sticky="ew")

        self.compactButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nCompaction",
                                                       command=self.forceToggle, onvalue=1, offvalue=0)
        self.compactButton.grid(row=2, column=0, padx=(10, 10), pady=(30, 30))

        self.forceEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="N", state="disabled",
                                                 validate='focusout', validatecommand=self.checkForceEntry, width=50)
        self.forceEntry.grid(row=2, column=3, padx=(10, 10), pady=(30, 30), sticky='w')

        self.forceLabel = customtkinter.CTkLabel(self.slide_frame, text="N    ")
        self.forceLabel.grid(row=2, column=4, padx=(0, 10), pady=(30, 30), sticky='e')

        ## Rotation Slider ##
        # Subsystem on/off control, slider and text entry for rotation oscillation frequency selection
        # Follows the same setup as the frequency slider section above
        self.rotate_var = "disabled"
        self.rotateSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                    number_of_steps=250, from_=0, to=50)
        self.rotateSlider.bind("<ButtonRelease-1>", self.checkRotateSlider)
        self.rotateSlider.grid(row=3, column=1, columnspan=2, padx=(10, 10), pady=(30, 30), sticky="ew")

        self.rotateButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nRotation",
                                                      command=self.rotateToggle, onvalue=1, offvalue=0)
        self.rotateButton.grid(row=3, column=0, padx=(10, 10), pady=(30, 30))

        self.rotateEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="Hz", state="disabled",
                                                  validate='focusout', validatecommand=self.checkRotateEntry, width=50)
        self.rotateEntry.grid(row=3, column=3, padx=(10, 10), pady=(30, 30), sticky='w')

        self.rotateLabel = customtkinter.CTkLabel(self.slide_frame, text="Hz")
        self.rotateLabel.grid(row=3, column=4, padx=(0, 10), pady=(30, 30), sticky='e')

        ## Run Time ##
        # Subsystem on/off control, slider and text entry for run time selection
        # Follows the same setup as the frequency slider section above
        self.time_var = "disabled"
        self.timeSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                  number_of_steps=250, from_=timeBounds[0], to=timeBounds[1])
        self.timeSlider.bind("<ButtonRelease-1>", self.checkTimeSlider)
        self.timeSlider.grid(row=4, column=1, columnspan=2, padx=(10, 10), pady=(30, 20), sticky="ew")

        self.timeButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nRun Time", command=self.timeToggle,
                                                    onvalue=1, offvalue=0)
        self.timeButton.grid(row=4, column=0, padx=(10, 10), pady=(30, 20))

        self.timeEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="sec.", state="disabled",
                                                validate='focusout', validatecommand=self.checkTimeEntry, width=50)
        self.timeEntry.grid(row=4, column=3, padx=(10, 10), pady=(30, 20), sticky='w')

        self.timeLabel = customtkinter.CTkLabel(self.slide_frame, text="sec. ")
        self.timeLabel.grid(row=4, column=4, padx=(0, 10), pady=(30, 20), sticky='e')

        # File Name Frame
        # Creates an entry box that allows for a desired filename to be entered. Upon test completion, all data will be saved
        # to an excel sheet under the entered name. Excel sheet formating can be found in "saveFile" function
        self.filename_frame = customtkinter.CTkFrame(self, width=100)                                                           # Create frame for fle widgets
        self.filename_frame.grid(row=0, column=1, rowspan=1, padx=(20, 0), pady=(20, 0), sticky="nsew")                         # Place the frame within the window
        self.results_var = tkinter.IntVar(value=0)

        self.filenameEntry = customtkinter.CTkEntry(master=self.filename_frame, placeholder_text='File_Name',                   # Create the text entry box
                                                    validate='focusout', width=200, height=30,                                  # File name will be checked once the user clicks outside of the entry box
                                                    validatecommand=self.checkFilename)                                         # Once the entry box loses focus, call function to validate the entered filename
        self.filenameEntry.grid(row=1, column=0, padx=10, pady=(5, 20), sticky='nsew', columnspan=2)                            # Place the entry box within the frame
        self.filenameLabel = customtkinter.CTkLabel(self.filename_frame, text='File Name', font=('CTkFont', 15))                # Create a label for the entry box
        self.filenameLabel.grid(row=0, column=0, padx=10, pady=(20, 5), sticky='sw')                                            # Place the label within the frame
        self.saveData_var = tkinter.IntVar(value=1)
        self.fileButtonNoSave = customtkinter.CTkRadioButton(master=self.filename_frame, text="Run Without Saving",             # Create radio button for the 'Do Not Save Data' option. Default option
                                                             variable=self.saveData_var, value=1, command=self.saveFile)        # Updates the 'saveFile' function with 'variable' and 'value' to control whether data is saved
        self.fileButtonNoSave.grid(row=2, column=0, padx=10, pady=(10, 20), sticky='nw')                                        # Place the radiobutton within the frame
        self.fileButtonSave = customtkinter.CTkRadioButton(master=self.filename_frame, text="Save Data",                        # Create radio button for the 'Do Save Data' option
                                                           variable=self.saveData_var, value=0, command=self.saveFile)          # Updates the 'saveFile' function with 'variable' and 'value' to control whether data is saved
        self.fileButtonSave.grid(row=2, column=1, padx=15, pady=(10, 20), sticky='nw')                                          # Place the radio button within the frame

        ## Program Info Frame ##
        # Creates a textbox that displays the directions on how to use the program
        self.info_frame = customtkinter.CTkFrame(self)                                                                          # Create frame for the textbox
        self.info_frame.grid(row=1, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")                                        # Place the frame within the window


        # Print Program Info Using Text Box
        # Create the text to be displayed within the textbox
        programInfo = '1) Select Cam amplitude (in mm)\n' \
                      '     *This will automatically set bounds for allowable frequencies\n\n' \
                      "2) Select which subsystems to enable using checkboxes\n" \
                      "     Vibration: Enables motor to run at desired frequency (in Hz),\n" \
                      "                        inducing vibration of the module\n" \
                      "     Compaction: Enables pneumatic piston to apply vertical \n" \
                      "                              compression force (in N) to the generant\n" \
                      "     Rotation: Enables servo motor to rotate the module in a \n" \
                      "                      fixed direction\n" \
                      "                      *Adjusting the rotation frequency (in Hz) will change the \n" \
                      "                       rotation direction at the desired rate\n" \
                      "     Run Time: Enables the user to select a desired run time (in sec.)\n" \
                      "                         *With 'Run Time' disabled, the system will run until the \n" \
                      "                          'STOP' button is selected\n\n" \
                      "3) Select if the user would like to save the data\n" \
                      "      *If data is to be saved, enter a filename and measured results \n" \
                      "        will be saved as aN .xlsx file\n" \
                      "      *If data is not to be saved, measured results will be displayed \n" \
                      "        without saving and filename may be ignored\n\n" \
                      "4) Once desired variables are entered, select the 'RUN' button to run \n" \
                      "      the system. \n" \
                      "      *Once the desired time is reached,\n" \
                      "        or the 'STOP' button is selected, the measured results will be \n" \
                      "        displayed and the data will be saved, if enabled\n"

        self.infoFrameText = customtkinter.CTkTextbox(master=self.info_frame, height=250)                                                   # Create the textbox
        self.infoFrameText.pack(fill='both', padx=10, pady=5)                                                                               # Place the textbox within the frame using 'pack' instead of grid. Easier method since there is only one widget
        self.infoFrameText.insert(0.0, programInfo)                                                                                         # Insert the desired text into the textbox
        self.infoFrameText.configure(state='disabled')                                                                                      # Disable the textbox so that the directions cannot be altered

        ## Run/Stop Button ##
        # Create the start/stop buttons for the system. Once the 'RUN' button is pressed, the desired subsystem parameters
        # will be sent through Serial Communication to the clearcore/VFD and the test will begin. The test will automatically
        # end if the run time exceeds the desired time. If the 'STOP' button is pressed while the test is running, the VFD will
        # be stopped and a command will be sent to the ClearCore to halt all other subsystems.
        self.button_frame = customtkinter.CTkFrame(self, fg_color='transparent')                                                            # Create the frame forthe start/stop buttons
        self.button_frame.grid(row=3, column=1, columnspan=1, padx=10, pady=10)                                                             # Place the frame within the window
        self.runButton = customtkinter.CTkButton(master=self.button_frame, width=225, height=75, corner_radius=25, fg_color='green',        # Create the 'START' button
                                                 border_color='#006400', hover_color='#228B22', border_width=5, font=('CTkFont', 20),       # Alter the appearence of the button
                                                 text='RUN', command=self.runButtonFunc)                                                    # Bind the button with the function that initiates the system start
                                                                                                                                            # 'RUN' button will be disabled until the test stops
        self.runButton.grid(row=3, column=1, padx=(25, 10), pady=0, sticky='sw')                                                            # Place the button within the frame

        self.stopButton = customtkinter.CTkButton(master=self.button_frame, width=225, height=75, corner_radius=25, fg_color='#EE2C2C',     # Create the 'STOP' button
                                                  border_color='#B22222', hover_color='#FF3030', border_width=5, font=('CTkFont', 20),      # Alter the appearence of the button
                                                  text='STOP', command=self.stopButtonFunc, state='disabled')                               # Bind the button with the function that initiates the system stop
                                                                                                                                            # 'STOP' button begins disabled until the 'RUN' button is pressed
        self.stopButton.grid(row=3, column=2, padx=10, pady=0, sticky='se')                                                                 # Place the button within the frame

        self.resultsWindow = None
    ## Functions to set bounds due to amplitude ##
    # Function checks the cam amplitude radio buttons and adjusts the parameter bounds
    # to safe ranges
    def setBounds(self):
        global freqBounds, forceBounds, rotSpeedBounds
        radio_var = self.radio_var.get()
        self.freqEntry.delete(0, len(self.freqEntry.get()))                     # Clear all subsystem parameter text entry boxes
        self.forceEntry.delete(0, len(self.forceEntry.get()))
        self.rotateEntry.delete(0, len(self.rotateEntry.get()))
        self.timeEntry.delete(0, len(self.timeEntry.get()))
        if radio_var == 0:                                                      # Check which radio button is selected
            amp = '1mm'
            freqBounds = [freqBounds1[0], freqBounds1[1]]                       # Assign the predetermined parameter bounds
            forceBounds = [forceBounds1[0], forceBounds1[1]]
            rotSpeedBounds = [rotSpeedBounds1[0], rotSpeedBounds1[1]]
        elif radio_var == 1:
            amp = '0.6mm'
            freqBounds = [freqBounds06[0], freqBounds06[1]]
            forceBounds = [forceBounds06[0], forceBounds06[1]]
            rotSpeedBounds = [rotSpeedBounds06[0], rotSpeedBounds06[1]]
        elif radio_var == 2:
            amp = '0.4mm'
            freqBounds = [freqBounds04[0], freqBounds04[1]]
            forceBounds = [forceBounds04[0], forceBounds04[1]]
            rotSpeedBounds = [rotSpeedBounds04[0], rotSpeedBounds04[1]]
        elif radio_var == 3:
            amp = 'custom'
            freqBounds = [freqBoundsCustom[0], freqBoundsCustom[1]]
            forceBounds = [forceBoundsCustom[0], forceBoundsCustom[1]]
            rotSpeedBounds = [rotSpeedBoundsCustom[0], rotSpeedBoundsCustom[1]]
            print("Custom cam amplitude selected\n*Make changes to allow for amplitude entry*")
        else:
            print('Something wrong with RadioButton function setBounds')
        self.freqSlider.configure(number_of_steps=250, from_=freqBounds[0], to=freqBounds[1])               # Apply the parameter bounds to the subsystems sliders
        self.forceSlider.configure(number_of_steps=250, from_=forceBounds[0], to=forceBounds[1])
        self.rotateSlider.configure(number_of_steps=250, from_=rotSpeedBounds[0], to=rotSpeedBounds[1])
        print('Subsystem bounds for ' + amp + ' amplitude')
        print('freqBounds:[' + str(freqBounds[0]) + ', ' + str(freqBounds[1]) + ']')
        print('forceBounds:[' + str(forceBounds[0]) + ', ' + str(forceBounds[1]) + ']')
        print('rotSpeedBounds:[' + str(rotSpeedBounds[0]) + ', ' + str(rotSpeedBounds[1]) + ']\n')

## Save data to excel sheet ##
# Function saves the received test data to an excel sheet with the desired filename
    def saveFile(self):
        if self.saveData_var.get() == 0:                                # If the data is to be saved
            fileNameValid = self.checkFilename()                            # Call function to determine if the entered filename is valid (proper formating, allowable characters, etc.)
            if fileNameValid:                                           # If the filename is valid
                fileName = self.filenameEntry.get()                     # Create the excel sheet with the desired filename
                wb = Workbook()
                wb.save(str(fileName)+'.xlsx')
                sheet = wb.active

                #Initialize sheet labels
                sheet.row_dimensions[2].height = 27                     # Size certain cells to allow for the labels
                sheet.column_dimensions['C'].width = 11
                sheet.column_dimensions['E'].width = 10

                sheet['D2'].value = 'Settings'                          # Add appropriate data labels within the excel sheet
                sheet['E2'] = 'Max\nMeasured'
                sheet['E2'].alignment = Alignment(wrapText=True)
                sheet['C3'].value = 'Freq.'
                sheet['C4'].value = 'Force'
                sheet['C5'].value = 'Rot. Freq'
                sheet['C6'].value = 'Time'
                sheet['F3'].value = 'Hz'
                sheet['F4'].value = 'N'
                sheet['F5'].value = 'Hz'
                sheet['F6'].value = 'Sec'
                sheet['C8'].value = 'Freq.'
                sheet['C9'].value = 'Force'
                sheet['C10'].value = 'Rot. Speed'
                sheet['C11'].value = 'Time'

                #Insert values into sheet
                sheet['D3'].value = float(self.freqEntry.get())         # Get the desired subsystems parameters and enter them into the excel sheet
                sheet['D4'].value = float(self.forceEntry.get())
                sheet['D5'].value = float(self.rotateEntry.get())
                sheet['D6'].value = float(self.timeEntry.get())
                sheet['E3'].value = maxFreq                             # Enter the max measured subsystem values into the excel sheet
                sheet['E4'].value = maxForce
                sheet['E5'].value = maxRotation
                sheet['E6'].value = maxTime

                for i in range(len(freqMeasures)):                                  # Enter the entirety of the measured data into the excel sheet
                    sheet.cell(row=8, column=i+4).value = freqMeasures[i]

                for i in range(len(forceMeasures)):
                    sheet.cell(row=9, column=i+4).value = forceMeasures[i]

                for i in range(len(rotationMeasures)):
                    sheet.cell(row=10, column=i+4).value = rotationMeasures[i]

                for i in range(len(t)):
                    sheet.cell(row=11, column=i+4).value = t[i]
                wb.save(str(fileName)+'.xlsx')                                      # Save the excel sheet
            else:
                print('Invalid filename')                                           # Return an error if the filename is invalid ********************************************
        elif self.saveData_var == 1:                                                # If the data is not to be saved skip the excel sheet creation
            print("Not Saving Data")

        else:                                                                       # Safety for if something goes wrong
            print('Error with save data radio buttons')


    ### Functions to disable sliders and entry boxes when related function is not enabled by the user ###

    def freqToggle(self):

        if self.vibButton.get() == 1:
            self.freqSlider.configure(state="normal")
            self.freqEntry.configure(state="normal")
        else:
            self.freqSlider.configure(state="disabled")
            self.freqEntry.delete(0, len(self.freqEntry.get()))
            self.freqEntry.configure(state="disabled")

    def forceToggle(self):
        if self.compactButton.get() == 1:
            self.forceSlider.configure(state="normal")
            self.forceEntry.configure(state="normal")
        else:
            self.forceSlider.configure(state="disabled")
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            self.forceEntry.configure(state="disabled")

    def rotateToggle(self):
        if self.rotateButton.get() == 1:
            self.rotateSlider.configure(state="normal")
            self.rotateEntry.configure(state="normal")
        else:
            self.rotateSlider.configure(state="disabled")
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            self.rotateEntry.configure(state="disabled")

    def timeToggle(self):
        if self.timeButton.get() == 1:
            self.timeSlider.configure(state="normal")
            self.timeEntry.configure(state="normal")
        else:
            self.timeSlider.configure(state="disabled")
            self.timeEntry.delete(0, len(self.timeEntry.get()))
            self.timeEntry.configure(state="disable")

    ### Functions to sync slider and entry values ###

    def checkFreqEntry(self):                                       # Check if entry is numeric (int or float)
        try:                                                        # First check if entry if integer
            val = int(self.freqEntry.get())
            val1 = True
        except ValueError:                                          # If integer check returns an error
            try:                                                    # Check if entry is a float
                val = float(self.freqEntry.get())
                val1 = True
            except ValueError:                                      # If both integer and float check return error
                val1 = False                                        # Set value to 'False' (indicates that entry is not numeric)

        if val1 == True:                                            # If entry is numeric
            if freqBounds[0] <= val <= freqBounds[1]:               # And if entry is within the bounds
                self.freqSlider.set(val)                            # Set the slider value to that of the entry box
                return True
        else:                                                       # If entry is not numeric or not within the bounds, delete the entry and print an error with the proper bounds
            self.freqEntry.delete(0, len(self.freqEntry.get()))     # Delete the entry
            print('Invalid frequency')
            print('Enter a value between ' + str(freqBounds[0]) + ' and ' + str(freqBounds[1]) + 'Hz')
            return False

    def checkFreqSlider(self, second):                                  # When the Slider is changed, update the Entry box to match
        if isinstance(self.freqEntry.get(), int):                       # Determine if existing entry is an integer
            if self.freqSlider.get() != int(self.freqEntry.get()):      # Determine if existing entry is equal to slider value
                self.freqEntry.delete(0, len(self.freqEntry.get()))     # If entry does not match the slider, delete the entry
                self.freqEntry.insert(self.freqSlider.get)              # Change entry box to the slider value
                return True
        else:
            self.freqEntry.delete(0, len(self.freqEntry.get()))         # If the entry value is not an integer, delete the entry
            self.freqEntry.insert(0, round(self.freqSlider.get(), 1))   # Change entry box to the slider value
            return True

# All following entry and slider check functions follow the formating of the frequency checks above
    def checkForceEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.forceEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.forceEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if forceBounds[0] <= val <= forceBounds[1]:
                self.forceSlider.set(val)
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            print('Invalid compaction force')
            print('Enter a value between ' + str(forceBounds[0]) + ' and ' + str(forceBounds[1]) + 'N')
            return False

    def checkForceSlider(self, second):                              #When the forceSlider is changed, update the forceEntry box to match
        if isinstance(self.forceEntry.get(), int):
            if self.forceSlider.get() != int(self.forceEntry.get()):
                self.forceEntry.delete(0, len(self.forceEntry.get()))
                self.forceEntry.insert(self.forceSlider.get)
                return True
        else:
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            self.forceEntry.insert(0, round(self.forceSlider.get(), 1))
            return True

    def checkRotateEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.rotateEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.rotateEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if rotSpeedBounds[0] <= val <= rotSpeedBounds[1]:
                self.rotateSlider.set(val)
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            print('Invalid rotation speed')
            print('Enter a value between ' + str(rotSpeedBounds[0]) + ' and ' + str(rotSpeedBounds[1]) + 'Hz')
            return False

    def checkRotateSlider(self, second):                              #When the freqSlider is changed, update the freqEntry box to match
        if isinstance(self.rotateEntry.get(), int):
            if self.rotateSlider.get() != int(self.rotateEntry.get()):
                self.rotateEntry.delete(0, len(self.rotateEntry.get()))
                self.rotateEntry.insert(self.rotateSlider.get)
                return True
        else:
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            self.rotateEntry.insert(0, round(self.rotateSlider.get(), 1))
            return True

    def checkTimeEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.timeEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.timeEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if timeBounds[0] <= val <= timeBounds[1]:   
                self.timeSlider.set(val)    
                return True
            elif val >= 0:
                self.timeSlider.configure(to=val)   #change slider from_ to
                self.timeSlider.set(val)            #set slider value to entry
                return True
        else:                                                       #If entry is not numeric or out of timeBounds, print an error with the proper bounds
            self.timeEntry.delete(0, len(self.timeEntry.get()))
            print('Invalid run time')
            print('Enter a value between ' + str(timeBounds[0]) + ' and ' + str(timeBounds[1]) + 'sec.')
            return False

    def checkTimeSlider(self, second):                              #When the freqSlider is changed, update the freqEntry box to match
        if isinstance(self.timeEntry.get(), int):
            if self.timeSlider.get() != int(self.timeEntry.get()):
                self.timeEntry.delete(0, len(self.timeEntry.get()))
                self.timeEntry.insert(self.timeSlider.get)
                return True
        else:
            self.timeEntry.delete(0, len(self.timeEntry.get()))
            self.timeEntry.insert(0, round(self.timeSlider.get(), 1))
            return True

## Validate the entered filename ##
# Function reads the entered filename if data is to be saved and determines if it is an acceptable name
    def checkFilename(self):
        filename = self.filenameEntry.get()                             # Get the entered filename
        if not re.findall(r'[^A-Za-z0-9_\-\\]', filename):              # Verify that the filename does not contain special characters
            if re.findall(r'[^0-9_\-\\]', filename[0]):
                fileNameValid = True
            else:
                fileNameValid = False
        else:
            fileNameValid = False
        return fileNameValid

## START function ##
# Function is called when the 'RUN' button is pressed
    def runButtonFunc(self):
        
        self.runButton.configure(state='disabled')                  # Disable the 'RUN' button
        self.stopButton.configure(state='normal', hover=True)       # Enable the 'STOP' button
        self.sendToClearcore()                                      # Call function to send parameters to the ClearCore
        #self.checkSerial()                                          # Calls function to check the serial port for return data

## STOP function ##
# Function is called when the 'STOP' button is pressed or 
# the program receives return data from ClearCore
    def stopButtonFunc(self):
        self.stopButton.configure(state='disabled')                 # Disable the 'STOP' button
        self.runButton.configure(state='normal', hover=True)        # Enable the 'RUN' button
        clearcore.write(str.encode("STOP:"))                         # Send the stop command to ClearCore. Pointless if test ended due to runTime, used if the 'STOP' button is pressed
        self.commandVFD(0)                                          # Command the VFD to stop
        self.finishTest()
        return

## Open new window to display the test results ##
    def openResultsWindow(self):
        if self.resultsWindow is None or not self.resultsWindow.winfo_exists():     # Open the 'ProcessResults' window if it is not already open
            self.resultsWindow = ProcessResults(self)
            self.resultsWindow.focus()                                              # Pull focus to the window

        else:
            self.resultsWindow.focus()

## Send to ClearCore ##
# Function is called when the 'RUN' button is pressed.
# Sends desired configuration settings to ClearCore through Serial Communication
    def sendToClearcore(self):

        if self.vibButton.get() == 1:                               # If the vibration subsystem is enabled
            setFreq = self.freqEntry.get()
            self.commandVFD(1, setFreq)
        else:
            self.commandVFD(0)                          # Set the frequency to the desired value

        if self.compactButton.get() == 1:                           # If the compaction subsystem is enabled
            setForce = self.forceEntry.get()                        # Set the compaction force to the desired value
        else:
            setForce = 999                                          # If compaction subsystem is disabled, this value will tell the ClearCore to leave the piston off

        if self.rotateButton.get() == 1:                            # If the rotation subsystem is enabled
            setRotation = self.rotateEntry.get()                    # Set the rotation oscillation frequency to the desired value
        else:
            setRotation = 999                                       # If the rotation subsystem is disabled, this value will tell the ClearCore to leave the rotation motor off

        if self.timeButton.get() == 1:                              # If the Run Time option is enabled
            setTime = self.timeEntry.get()                          # Set the run time to the desired value
        else:
            setTime = 999                                           # if the Run Time option is diabled, this value will tell the ClearCore to run the system until a 'STOP' command is received
        sendComm = 'Force:' + str(setForce) + ',Rotation:' + str(setRotation) + \
                   ',Time:' + str(setTime) + ',Go:'                                                             # String of desired subsyem values
        clearcore.write(str.encode(sendComm))
        global app
        app.after(10000, self.checkSerial())                                                                   # Send the string to ClearCore via Serial Communication

                                                                                                                # Freq doesn't need to be sent ***********************************


    def readData(self):
        global freqDataRecieved, forceDataRecieved, heightDataRecieved, timeDataRecieved
        recievedData = clearcore.readline().decode()
        print(recievedData)
        splitData = recievedData.split("\n")    #String format "Freq:12,23,34\nForce:12,23,34\n"
        for i in range(0, len(splitData)):
            text = splitData[i]
            text = text.split(":")

            if text[0] == "Freq":
                freqDataRecieved = text[1].split(",")
                for j in range(0, len(freqDataRecieved)):
                    freqDataRecieved[j] = float(freqDataRecieved[j])
            elif text[0] == "Force":
                forceDataRecieved = text[1].split(",")
                for j in range(0, len(forceDataRecieved)):
                    forceDataRecieved[j] = float(forceDataRecieved[j])
            elif text[0] == "Time":
                timeDataRecieved = text[1].split(",")
                for j in range(0, len(timeDataRecieved)):
                    timeDataRecieved[j] = float(timeDataRecieved[j])
            elif text[0] == "Height":
                heightDataRecieved[j] = text[1].split(",")
                for j in range(0, len(heightDataRecieved)):
                    heightDataRecieved[j] = float(heightDataRecieved[j])
            elif text[0] == "STOP":
                self.finishTest()

    def checkSerial(self):
        print("checkSerial")
        #print(clearcore.readline().decode())
        if clearcore.inWaiting() > 0:
                print("bits waiting")
                #print(clearcore.readline().decode())
                self.readData()
                
        
    
    def finishTest(self):
        self.stopButton.configure(state='disabled')
        self.runButton.configure(state='normal', hover=True)

        self.commandVFD(0)  #Turn off VFD

        self.readData()
        self.openResultsWindow()
        self.saveFile()

    def commandVFD(self, VFD_cmd, VFD_Hz=0):
        instrument = minimalmodbus.Instrument('COM5', 1)
        instrument.serial.port = 'COM5'
        instrument.serial.baudrate = 9600
        instrument.serial.bytesize = 8
        instrument.serial.parity = serial.PARITY_NONE
        instrument.serial.stopbits = 1
        instrument.serial.timeout = 1
        instrument.address = 1
        instrument.debug = False                                        #Turn on for debug mode
        instrument.mode = minimalmodbus.MODE_RTU
        VFD_Hz = float(VFD_Hz)

        if VFD_cmd == 1:
            instrument.write_register(8193, VFD_Hz, 2, 6)               #Set VFD frequency
            instrument.write_register(8192, 1, 0, 6)                    #Command to start VFD
        elif VFD_cmd == 0:
            instrument.write_register(8192, 8, 0, 6)                    #Command to stop the VFD

if __name__ == "__main__":
    app = App()
    app.mainloop()